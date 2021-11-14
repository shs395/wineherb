from openpyxl import load_workbook
from selenium import webdriver
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
from urllib.parse import unquote_plus
# from webdriver_manager.chrome import ChromeDriverManager
import time
import platform
import chromedriver_autoinstaller
import subprocess
from selenium.webdriver.chrome.options import Options

print("current platform : " + platform.system())

if(platform.system() == 'Windows'):
    try:          
        subprocess.Popen(r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\chrometemp"')
    except:
        subprocess.Popen(r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:/ChromeTEMP"')

    option = Options()
    option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
    try:
        driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=option)
    except:
        chromedriver_autoinstaller.install(True)
        driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=option)
    driver.implicitly_wait(10)

elif(platform.system() == 'Darwin'):
    subprocess.Popen(r'/Applications/Google\ Chrome.app/Contents/MacOS/Google\ Chrome --remote-debugging-port=9222 --user-data-dir="~/ChromeProfile"', shell=True)

    chromedriver_autoinstaller.install()

    co = Options()
    co.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
    driver = webdriver.Chrome(options=co)
    driver.implicitly_wait(10)

# 그냥 실행
# driver = webdriver.Chrome(ChromeDriverManager().install())

# 엑셀파일 불러오기
load_wb = load_workbook("./result.xlsx", data_only=True)

# Sheet 라는 이름의 시트 선택
ws1 = load_wb['Sheet']

# 검색을 원하는 아이템(와인 이름)과 사이트에 따라 검색하여 첫번째 사이트의 url을 알려줌
def getFirstSiteUrl(item, site):
    try:
        html = searchGoogle(item, site)
        soup = BeautifulSoup(html, 'html.parser')
        # type(siteUrlList) == <class 'bs4.element.ResultSet'>
        siteUrlList = soup.select('.yuRUbf')
        firstSiteUrl = siteUrlList[0].a.attrs['href']
        return firstSiteUrl
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# 해당 site에서 관련 item 찾는 구글 검색
def searchGoogle(item, site):
    try:
        # 구글 검색을 위한 기본 url
        baseUrl = 'https://www.google.com/search?q='
        # 검색을 위한 최종 url
        url = baseUrl + quote_plus(item+site)
        driver.get(url)
        # return 값 : 검색을 통해 나온 html
        return driver.page_source
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# wine.com 에서 레이팅 가져오기
def getRatingInWineCom(url):
    try:
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # type(ratings) == <class 'bs4.element.ResultSet'>
        ratings = soup.select('div.prodInfo > ul.wineRatings_list > li.wineRatings_listItem')
        rating = ''
        for i in ratings:
            rating = rating + i.select_one('.wineRatings_initials').text + ' ' + i.select_one('.wineRatings_rating').text + ' '
        return rating
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# winesearcher 에서 평균 가격 가져오기
def getAvgPriceInWineSearcher(url):
    try:
        url = url + unquote_plus('?Xcurrencycode=USD&Xsavecurrency=Y')
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        avgPrice = soup.select_one('div.smaller > strong.text-nowrap').text
        avgPrice = avgPrice.replace(" ", "")
        avgPrice = avgPrice.replace("\n", "")
        return avgPrice
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

def getGrapeInWineSearcher(url):
    url = url + unquote_plus('#t2')
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    spanList = soup.select('div#tab-info > div.card-columns > div')
    #div#tab-info > div.card-columns > div > div.card-body > div > a > span.font-ligjt-bold
    print(spanList)
    # firstSiteUrl = siteUrlList[0].a.attrs['href']


def getAlcholInWineSearcher(url):
    url = url + unquote_plus('#t2')
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    
# max_row = 값이 있는 최대 행렬의 줄 값
# 2번 부터 최대 행렬까지
for row in range(2, ws1.max_row+1):
    # 값이 없으면 그만하기
    if(ws1.cell(row,1).value is None):
        break

    else:
        # 검색을 원하는 와인 이름
        wineName = ws1.cell(row,1).value

        # column2 = wine.com 주소
        wineComFirstSiteUrl = getFirstSiteUrl(wineName, ' WINE.COM')
        ws1.cell(row=row, column= 2).value = wineComFirstSiteUrl
        load_wb.save('result.xlsx')

        # column3 = wine.com 레이팅
        rating = getRatingInWineCom(wineComFirstSiteUrl)
        ws1.cell(row=row, column= 3).value = rating
        load_wb.save('result.xlsx')

        # column4 = vivino 주소
        vivinoFirstSiteUrl = getFirstSiteUrl(wineName, ' vivino')
        ws1.cell(row=row, column= 4).value = vivinoFirstSiteUrl
        load_wb.save('result.xlsx')
        
        # column5 = wine-searcher 주소
        wineSearcherFirstSiteUrl = getFirstSiteUrl(wineName, ' winesearcher')
        ws1.cell(row=row, column= 5).value = wineSearcherFirstSiteUrl 
        load_wb.save('result.xlsx')

        # column6 = wine-searcher 평균 가격
        avgPrice = getAvgPriceInWineSearcher(wineSearcherFirstSiteUrl)
        ws1.cell(row=row, column= 6).value = avgPrice
        load_wb.save('result.xlsx')

        # getGrapeInWineSearcher(wineSearcherFirstSiteUrl)
    
    time.sleep(3)
