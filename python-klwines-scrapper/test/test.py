from openpyxl import load_workbook
from selenium import webdriver
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
from urllib.parse import unquote_plus
# from webdriver_manager.chrome import ChromeDriverManager
import time
import platform
import chromedriver_autoinstaller
import subprocess
from selenium.webdriver.chrome.options import Options

print("current platform : " + platform.system())

if(platform.system() == 'Windows'):
    try:          
        subprocess.Popen(r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\chrometemp"')
    except:
        subprocess.Popen(r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:/ChromeTEMP"')

    option = Options()
    option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
    try:
        driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=option)
    except:
        chromedriver_autoinstaller.install(True)
        driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=option)
    driver.implicitly_wait(10)

elif(platform.system() == 'Darwin'):
    subprocess.Popen(r'/Applications/Google\ Chrome.app/Contents/MacOS/Google\ Chrome --remote-debugging-port=9222 --user-data-dir="~/ChromeProfile"', shell=True)

    chromedriver_autoinstaller.install()

    co = Options()
    co.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
    driver = webdriver.Chrome(options=co)
    driver.implicitly_wait(10)

# klwines 계정정보 불러오기
with open("../../klwines_account.txt") as f:
    lines = f.readlines()
    klwines_id = lines[0].strip()
    klwines_pw = lines[1].strip()

# 엑셀파일 불러오기
load_wb = load_workbook("./klwines.xlsx", data_only=True)

# Sheet 라는 이름의 시트 선택
ws1 = load_wb['Sheet']

# klwines 로그인
def loginKlwines(id, pw):
    # 로그인 안되어 있는 경우
    try:
        url = 'https://www.klwines.com/account/login'
        driver.get(url)
        driver.find_element_by_id('Email').send_keys(id)
        driver.find_element_by_id('Password').send_keys(pw)
        driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/div/div[2]/form/div[1]/table/tbody/tr[3]/td/input').click()
    # 로그인 정보가 남아 로그인 되어 있는 경우
    except:
        pass

# 해당 site에서 관련 item 찾는 klwines에서 검색
def searchKlwines(item):
    try:
        # 구글 검색을 위한 기본 url
        baseUrl = 'https://www.klwines.com/Products?searchText='
        # 검색을 위한 최종 url
        url = baseUrl + quote_plus(item)
        driver.get(url)
        # return 값 : 검색을 통해 나온 html
        return driver.page_source
    except Exception as e:
        print('서치 에러', e)
        return 'ERROR'

def isLastPage(html):
    try:
        soup = BeautifulSoup(html, 'html.parser')
        nextButton = soup.select('.page-filters-block > .floatLeft > a')[-1].text
        if(nextButton == 'next'):
            return False
        else:
            return True
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'  

def printPresentPage(html):
    try:
        soup = BeautifulSoup(html, 'html.parser')
        presentPage = soup.select_one('.page-filters-block > .floatLeft > strong').text
        print('현재 페이지', presentPage)
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    
def goNextPage(html):
    try:
        soup = BeautifulSoup(html, 'html.parser')
        nextPageLink = soup.select('.page-filters-block > .floatLeft > a')[-1].attrs['href']
        baseUrl = 'https://www.klwines.com'
        url = baseUrl + unquote_plus(nextPageLink)
        driver.get(url)
        return driver.page_source
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    
def getWineList(html):
    try:
        soup = BeautifulSoup(html, 'html.parser')
        wineList = soup.select('.tf-product')
        return wineList
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    

# button 타이틀 가져오기, Add To Cart or Place Bid
def isAddToCart(item):
    try:
        if(item.select('.tf-button button')[0].attrs['title'] == 'Add To Cart'):
            return True
        else:
            return False
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    

def getWineName(wineInfo):
    try:
        return wineInfo.select_one('.tf-product-header > a').text.replace('\n', '').strip()
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

def getWineImgLink(wineInfo):
    try:
        return wineInfo.select_one('.tf-product-image > a > img').attrs['src']
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

def getWinePrice(wineInfo):
    try:
        return wineInfo.select_one('.tf-price > .global-pop-color').text
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

def getWineRating(wineInfo):
    try:
        rating = ''
        wineRating = wineInfo.select('.tf-pill-container > a')
        for i in wineRating:
                rating = rating + i.text.replace(':', '').replace('\n', '') + ' '
        return rating
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    
def getWineQuantityUrl(wineInfo):
    try:
        wineQuantityLink = wineInfo.select_one('.tf-product-header').a.attrs['href']
        baseUrl = 'https://www.klwines.com'
        url = baseUrl + unquote_plus(wineQuantityLink)
        return url
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    
def getWineQuantity(wineInfo):
    try:
        url = getWineQuantityUrl(wineInfo)
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        wineQuantityInfo = soup.select('.inventory > div.column > table > tbody > tr > td')
        Quantity = ''
        for i in wineQuantityInfo:
            Quantity = Quantity + i.text.replace(' ', '').replace('\n', '') + ' '
        return Quantity
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'
    
def getWineInfos(html, row):
    # 와인리스트 가져오기 ype(wineList) == <class 'bs4.element.ResultSet'> len=50
    wineList = getWineList(html)
    for i in range(len(wineList)):
        if(isAddToCart(wineList[i]) == True):
            
            # column1 = 상품명
            wineName = getWineName(wineList[i])
            ws1.cell(row=row, column= 1).value = wineName
            load_wb.save('klwines.xlsx')

            # column2 = 가격
            winePrice = getWinePrice(wineList[i])
            ws1.cell(row=row, column= 2).value = winePrice
            load_wb.save('klwines.xlsx')
            # column3 = 평점
            wineRating = getWineRating(wineList[i])
            ws1.cell(row=row, column= 3).value = wineRating
            load_wb.save('klwines.xlsx')

            # column4 = 재고수
            wineQuantity = getWineQuantity(wineList[i])
            ws1.cell(row=row, column= 4).value = wineQuantity
            load_wb.save('klwines.xlsx')
            time.sleep(0.5)

            # column5 = 이미지 url
            wineImgLink = getWineImgLink(wineList[i])
            ws1.cell(row=row, column= 5).value = wineImgLink
            load_wb.save('klwines.xlsx')
            
            row = row + 1
        else:
            pass
    
    return row


# 검색어 가져오기
search_keyword = ws1.cell(1,7).value
print(search_keyword)

# klwines 로그인
loginKlwines(klwines_id, klwines_pw)

# klwines 검색
html = searchKlwines(search_keyword)
row = 2
while(1):
    # 마지막 페이지인경우
    if(isLastPage(html) == True):
        printPresentPage(html)
        row = getWineInfos(html, row)
        break
    # 페이지가 더 있는 경우
    else:
        #현재 페이지 확인
        printPresentPage(html)
        row = getWineInfos(html, row)
        html = goNextPage(html)
