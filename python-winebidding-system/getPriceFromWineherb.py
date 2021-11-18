from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
from urllib.parse import unquote_plus
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
from functions import login
from functions import openChrome


def getWineNameFromWineBid(url):
    try:
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        wineName = soup.select_one('div.item > h1').text
        return wineName
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# 검색을 원하는 아이템(와인 이름)과 사이트에 따라 검색하여 첫번째 사이트의 url을 알려줌
def getFirstSiteUrl(item, site):
    try:
        html = searchGoogle(item, site)
        soup = BeautifulSoup(html, 'html.parser')
        # type(siteUrlList) == <class 'bs4.element.ResultSet'>
        siteUrlList = soup.select('.yuRUbf')
        firstSiteUrl = siteUrlList[0].a.attrs['href']
        return firstSiteUrl
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# 해당 site에서 관련 item 찾는 구글 검색
def searchGoogle(item, site):
    try:
        # 구글 검색을 위한 기본 url
        baseUrl = 'https://www.google.com/search?q='
        # 검색을 위한 최종 url
        url = baseUrl + quote_plus(item+site)
        driver.get(url)
        # return 값 : 검색을 통해 나온 html
        return driver.page_source
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

def getPricePageInWineSearcher(url):
    try:
        url = url + unquote_plus('/-/-/ndb?Xbottle_size=Bottle')
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        return soup
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# winesearcher 에서 평균 가격 가져오기
def getAvgPriceInWineSearcher(soup):
    try:
        avgPrice = soup.select_one('div.smaller > strong.text-nowrap').text
        avgPrice = avgPrice.replace(" ", "")
        avgPrice = avgPrice.replace("\n", "")
        return avgPrice
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# winesearcher 에서 제일 낮은 가격 가져오기
def getLowestPriceInWineSearcher(soup):
    try:
        lowestPriceInteger = soup.select('span.price__integer-part')[0].text
        lowestPriceFractional = soup.select('span.price__fractional-part')[0].text
        lowestPrice = lowestPriceInteger + lowestPriceFractional
        return lowestPrice
    except Exception as e:
        print('에러 발생', e)
        return 'ERROR'

# 엑셀파일 불러오기
load_wb = load_workbook("./wineAutoBid.xlsx", data_only=True)

# Sheet 라는 이름의 시트 선택
ws1 = load_wb['Sheet1']

# selenium 실행 및 driver 정보 가져오기
driver = openChrome.openChrome()

# max_row = 값이 있는 최대 행렬의 줄 값
# 2번 부터 최대 행렬까지
for row in range(2, ws1.max_row+1):
    # 값이 없으면 그만하기
    if(ws1.cell(row,1).value is None):
        break

    else:
        # 검색을 원하는 winebid 주소
        wineBidUrl = ws1.cell(row,1).value

        # column2 = wine 이름
        wineName = getWineNameFromWineBid(wineBidUrl)
        ws1.cell(row=row, column= 2).value = wineName
        load_wb.save('./wineAutoBid.xlsx')

        # column3 = wine-searcher 주소
        wineSearcherFirstSiteUrl = getFirstSiteUrl(wineName, ' winesearcher')
        ws1.cell(row=row, column= 3).value = wineSearcherFirstSiteUrl 
        load_wb.save('./wineAutoBid.xlsx')

        soup = getPricePageInWineSearcher(wineSearcherFirstSiteUrl)

        # column4 = winesearcher avg price
        avgPrice = getAvgPriceInWineSearcher(soup)
        ws1.cell(row=row, column= 4).value = avgPrice
        load_wb.save('./wineAutoBid.xlsx')

        # column5 = winesearcher lowest price
        lowestPrice = getLowestPriceInWineSearcher(soup)
        ws1.cell(row=row, column= 5).value = lowestPrice
        load_wb.save('./wineAutoBid.xlsx')


